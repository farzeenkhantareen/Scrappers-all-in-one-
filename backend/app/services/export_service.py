"""
Export service: generates CSV, Excel, JSON, XML, PDF, ZIP exports.
"""

import csv
import io
import json
import os
import uuid
import zipfile
from datetime import datetime
from typing import Any

from app.config import settings


class ExportService:
    """Handles all data export formats."""

    @staticmethod
    def export_json(data: Any, filename: str = None) -> str:
        """Export data as JSON file and return file path."""
        if not filename:
            filename = f"export_{uuid.uuid4().hex[:8]}.json"
        filepath = os.path.join(settings.EXPORT_DIR, filename)
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, default=str, ensure_ascii=False)
        return filepath

    @staticmethod
    def export_csv(rows: list[dict], filename: str = None) -> str:
        """Export list of dicts as CSV."""
        if not filename:
            filename = f"export_{uuid.uuid4().hex[:8]}.csv"
        filepath = os.path.join(settings.EXPORT_DIR, filename)
        if not rows:
            with open(filepath, "w") as f:
                f.write("")
            return filepath

        fieldnames = list(rows[0].keys())
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                # Flatten nested dicts/lists to strings
                flat_row = {}
                for k, v in row.items():
                    if isinstance(v, (dict, list)):
                        flat_row[k] = json.dumps(v, ensure_ascii=False)
                    else:
                        flat_row[k] = v
                writer.writerow(flat_row)
        return filepath

    @staticmethod
    def export_excel(sheets: dict[str, list[dict]], filename: str = None) -> str:
        """Export multiple sheets as Excel .xlsx file."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise RuntimeError("openpyxl not installed")

        if not filename:
            filename = f"export_{uuid.uuid4().hex[:8]}.xlsx"
        filepath = os.path.join(settings.EXPORT_DIR, filename)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for sheet_name, rows in sheets.items():
            ws = wb.create_sheet(title=sheet_name[:31])
            if not rows:
                continue

            headers = list(rows[0].keys())
            # Header row styling
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1a1a2e")
                cell.alignment = Alignment(horizontal="center")

            # Data rows
            for row_idx, row in enumerate(rows, 2):
                for col_idx, key in enumerate(headers, 1):
                    val = row.get(key, "")
                    if isinstance(val, (dict, list)):
                        val = json.dumps(val, ensure_ascii=False)
                    ws.cell(row=row_idx, column=col_idx, value=val)

            # Auto-width columns
            for col in ws.columns:
                max_len = max(
                    (len(str(cell.value)) for cell in col if cell.value), default=10
                )
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        wb.save(filepath)
        return filepath

    @staticmethod
    def export_xml(data: Any, root_tag: str = "results", filename: str = None) -> str:
        """Export data as XML."""
        import xml.etree.ElementTree as ET

        if not filename:
            filename = f"export_{uuid.uuid4().hex[:8]}.xml"
        filepath = os.path.join(settings.EXPORT_DIR, filename)

        root = ET.Element(root_tag)

        def _add_to_xml(parent: ET.Element, data: Any, tag: str = "item"):
            if isinstance(data, dict):
                el = ET.SubElement(parent, tag)
                for k, v in data.items():
                    _add_to_xml(el, v, tag=str(k))
            elif isinstance(data, list):
                for item in data:
                    _add_to_xml(parent, item, tag=tag.rstrip("s"))
            else:
                parent.text = str(data) if data is not None else ""

        if isinstance(data, list):
            for item in data:
                _add_to_xml(root, item, "item")
        else:
            _add_to_xml(root, data, "data")

        tree = ET.ElementTree(root)
        ET.indent(tree, space="  ")
        tree.write(filepath, encoding="utf-8", xml_declaration=True)
        return filepath

    @staticmethod
    def export_pdf(data: dict, title: str = "Scraper Export", filename: str = None) -> str:
        """Export structured report as PDF."""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib import colors
        except ImportError:
            raise RuntimeError("reportlab not installed")

        if not filename:
            filename = f"export_{uuid.uuid4().hex[:8]}.pdf"
        filepath = os.path.join(settings.EXPORT_DIR, filename)

        doc = SimpleDocTemplate(filepath, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                                 topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph(title, styles["Title"]))
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
        story.append(Spacer(1, 0.5*cm))

        def _add_section(key: str, value: Any, level: int = 0):
            indent = "  " * level
            if isinstance(value, dict):
                story.append(Paragraph(f"{indent}<b>{key}</b>", styles["Heading3"]))
                for k, v in value.items():
                    _add_section(k, v, level + 1)
            elif isinstance(value, list):
                story.append(Paragraph(f"{indent}<b>{key}</b> ({len(value)} items)", styles["Heading3"]))
                if value and isinstance(value[0], dict):
                    # Render as table
                    headers = list(value[0].keys())[:6]
                    table_data = [headers] + [
                        [str(row.get(h, ""))[:50] for h in headers]
                        for row in value[:20]
                    ]
                    t = Table(table_data, repeatRows=1)
                    t.setStyle(TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ]))
                    story.append(t)
                    story.append(Spacer(1, 0.3*cm))
            else:
                story.append(Paragraph(f"{indent}<b>{key}:</b> {str(value)[:200]}", styles["Normal"]))

        for k, v in data.items():
            _add_section(k, v)
            story.append(Spacer(1, 0.3*cm))

        doc.build(story)
        return filepath

    @staticmethod
    def export_zip(file_paths: list[str], filename: str = None) -> str:
        """Bundle multiple files into a ZIP archive."""
        if not filename:
            filename = f"export_{uuid.uuid4().hex[:8]}.zip"
        filepath = os.path.join(settings.EXPORT_DIR, filename)

        with zipfile.ZipFile(filepath, "w", zipfile.ZIP_DEFLATED) as zf:
            for fp in file_paths:
                if os.path.exists(fp):
                    zf.write(fp, arcname=os.path.basename(fp))

        return filepath

    @staticmethod
    def get_file_size(filepath: str) -> int:
        """Get file size in bytes."""
        try:
            return os.path.getsize(filepath)
        except OSError:
            return 0
